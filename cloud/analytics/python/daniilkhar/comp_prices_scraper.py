import requests
import json
import re
from bs4 import BeautifulSoup
import pandas as pd
import time
import getpass
import yt
import yt.wrapper
from openpyxl import load_workbook
from datetime import date



user_agent = '(Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko)' \
             ' Chrome/76.0.3809.132 YaBrowser/19.9.2.273 Yowser/2.5 Safari/537.36'
VAT = 0.2


class GetPrices:

    def __init__(self):
        self.competitor_name = ''
        self.service = {'ssd': {'pricing_unit': 'gb*month', 'links': '', 'comments': '', 'price': {'on_demand': None}}, 
                        'hdd': {'pricing_unit': 'gb*month', 'links': '', 'comments': '', 'price': {'on_demand': None}}, 
                        'cold_storage': {'pricing_unit': 'gb*month', 'links': '', 'comments': '', 'price': {'on_demand': None}}, 
                        'cold_read': {'pricing_unit': '10000_requests', 'links': '', 'comments': '', 'price': {'on_demand': None}}, 
                        'cold_write': {'pricing_unit': '1000_requests', 'links': '', 'comments': '', 'price': {'on_demand': None}}, 
                        'standard_storage': {'pricing_unit': 'gb*month', 'links': '', 'comments': '', 'price': {'on_demand': None}}, 
                        'standard_read': {'pricing_unit': '10000_requests', 'links': '', 'comments': '', 'price': {'on_demand': None}}, 
                        'standard_write': {'pricing_unit': '1000_requests', 'links': '', 'comments': '', 'price': {'on_demand': None}},
                        'cpu': {'pricing_unit': 'core*month', 'links': '', 'comments': '', 
                                'price': {'on_demand': None, '1_year': None, '3_year': None}},
                        'ram': {'pricing_unit': 'gb*month', 'links': '', 'comments': '', 
                                'price': {'on_demand': None, '1_year': None, '3_year': None}},
                        'windows': {'pricing_unit': 'item*month', 'links': '', 'comments': '', 
                                    'price': {'on_demand': None, '1_year': None, '3_year': None}},
                        '2_4_windows': {'pricing_unit': 'item*month', 'links': '', 'comments': '', 
                                        'price': {'on_demand': None, '1_year': None, '3_year': None}},
                        '2_8_windows': {'pricing_unit': 'item*month', 'links': '', 'comments': '', 
                                        'price': {'on_demand': None, '1_year': None, '3_year': None}},
                        '2_4_linux': {'pricing_unit': 'item*month', 'links': '', 'comments': '', 
                                      'price': {'on_demand': None, '1_year': None, '3_year': None}},
                        '2_8_linux': {'pricing_unit': 'item*month', 'links': '', 'comments': '', 
                                      'price': {'on_demand': None, '1_year': None, '3_year': None}},
                        '2_4_80_windows': {'pricing_unit': 'item*month', 'links': '', 'comments': '', 
                                           'price': {'on_demand': None, '1_year': None, '3_year': None}},
                        '2_8_160_windows': {'pricing_unit': 'item*month', 'links': '', 'comments': '', 
                                            'price': {'on_demand': None, '1_year': None, '3_year': None}},
                        '2_4_80_linux': {'pricing_unit': 'item*month', 'links': '', 'comments': '', 
                                         'price': {'on_demand': None, '1_year': None, '3_year': None}},
                        '2_8_160_linux': {'pricing_unit': 'item*month', 'links': '', 'comments': '', 
                                          'price': {'on_demand': None, '1_year': None, '3_year': None}}
                       }
        self.currency = 'USD'
        self.region = ''

    def get_Azure_prices(self):
        self.__init__()
        self.competitor_name = "Azure"
        self.region = 'West Europe'
        self.currency = 'USD'
        self.service['ssd']['comments'] = 'P10'
        self.service['hdd']['comments'] = 'S10'
        self.service['cold_storage']['comments'] = 'First 50 TB'
        self.service['standard_storage']['comments'] = 'First 50 TB'
        self.service['2_4_linux']['comments'] = 'A2v2'
        self.service['2_8_linux']['comments'] = 'D2v3'
        self.service['2_4_windows']['comments'] = 'A2v2'
        self.service['2_8_windows']['comments'] = 'D2v3'
        self.service['2_4_80_windows']['comments'] = 'A2v2 + P10'
        self.service['2_8_160_windows']['comments'] = 'D2v3 + P10'
        self.service['2_4_80_linux']['comments'] = 'A2v2 + P10'
        self.service['2_8_160_linux']['comments'] = 'D2v3 + P10'
        for name in ['ssd', 'hdd']:
            self.service[name]['links'] = 'https://azure.microsoft.com/ru-ru/pricing/details/managed-disks/'
        for name in ['cold_storage', 'cold_read', 'cold_write', 'standard_storage', 'standard_read', 'standard_write']:
            self.service[name]['links'] = 'https://azure.microsoft.com/ru-ru/pricing/details/managed-disks/'
        for name in ['2_4_windows', '2_8_windows', '2_4_80_windows', '2_8_160_windows']:
            self.service[name]['links'] = 'https://azure.microsoft.com/ru-ru/pricing/details/virtual-machines/windows/'
        for name in ['2_4_linux', '2_8_linux', '2_4_80_linux', '2_8_160_linux']:
            self.service[name]['links'] = 'https://azure.microsoft.com/ru-ru/pricing/details/virtual-machines/windows/'

        # SSD HDD prices
        req_ssd_hdd = requests.get('https://azure.microsoft.com/ru-ru/pricing/details/managed-disks/',
                                   headers={'User-Agent': user_agent})
        soup_ssd_hdd = BeautifulSoup(req_ssd_hdd.text, "lxml")
        tables_ssd_hdd = soup_ssd_hdd.find_all('table', attrs={'class': 'sd-table'})
        for table in tables_ssd_hdd:
            for row in table.find_all('tr'):
                if row.find('strong'):
                    type_ = str(row.find('strong').text).replace(' ', '').replace('\n', '').replace('\r', '')
                    if type_ == 'P10':
                        disk_size = float(row.find_all('td')[1].text.split('\xa0')[0])
                        ssd_price_tb = json.loads(row.find('span')['data-amount'])['regional']['europe-west']
                        self.service['ssd']['price']['on_demand'] = ssd_price_tb/disk_size
                    if type_ == 'S10':
                        disk_size = float(row.find_all('td')[1].text.split('\xa0')[0])
                        hdd_price_tb = json.loads(row.find('span')['data-amount'])['regional']['europe-west']
                        self.service['hdd']['price']['on_demand'] = hdd_price_tb/disk_size

        # Cold Hot Storage Prices
        type_cold = ['cold_storage', 'cold_write', 'cold_read']
        type_standard = ['standard_storage', 'standard_write', 'standard_read']
        req_cold_hot = requests.get('https://azure.microsoft.com/ru-ru/pricing/details/storage/blobs/',
                                    headers={'User-Agent': user_agent})
        soup_cold_hot = BeautifulSoup(req_cold_hot.text, "lxml")
        section_cold_hot = soup_cold_hot.find('section', attrs={'data-filter': 'recommended-filter'})
        tables_cold_hot = section_cold_hot.find_all('div', attrs={'class': 'storage-table lrs'})
        for table in tables_cold_hot:
            for tbody in table.find_all('tbody'):
                for trow in tbody.find_all('tr'):
                    count = 0
                    flag = 0
                    for tdata in trow.find_all('td'):
                        #Finding the right row
                        if (tdata.text == '\r\nПервые 50\xa0ТБ в месяц                            ') \
                                and (count == 0):
                            flag = 1 # Storage prices
                        if (tdata.text == 'Операции записи (за 10\xa0000)1') \
                                and (count == 0):
                            flag = 2 # Write prices
                        if (tdata.text ==
                            '\n\n\nОперации чтения (за 10\xa0000)3\n\nОперации чтения данных из архива '
                            'с высоким приоритетом (за 10\xa0000)5\n\n\n\n') \
                                and (count == 0):
                            flag = 3 # Read prices
                        #Writing the prices
                        if flag in {1, 2, 3} and count in {2, 3}:
                            price = json.loads(tdata.find('span')['data-amount'])['regional']['europe-west']
                            if flag == 2:
                                price = price / 10
                            if count == 2:
                                self.service[type_standard[flag - 1]]['price']['on_demand'] = price
                            elif count == 3:
                                self.service[type_cold[flag - 1]]['price']['on_demand'] = price
                        count += 1

        # Standard Windows Configuration Prices
        req = requests.get('https://azure.microsoft.com/ru-ru/pricing/details/virtual-machines/windows/',
                                              headers={'User-Agent': user_agent})
        soup_win = BeautifulSoup(req.text, "lxml")
        tmp = soup_win.find('td', attrs={'class': 'webdirect-price av2-standard-series-ahb-switch'})
        tbody = tmp.find_parent('tbody')
        for row in tbody.find_all('tr'):
            data = row.find_all('td')
            if data[1].text.replace(' ', '').replace('\r\n', '').replace('\xa0', '') == 'A2v2':
                try:
                    price = float(json.loads(data[5].find('span')['data-amount'])['regional']['europe-west'])
                    self.service['2_4_windows']['price']['on_demand'] = price * 720 - self.service['ssd']['price']['on_demand'] * 20
                except Exception as err:
                    print('No data for Azure win A2 v2. ', err)

        tmp = soup_win.find('td', attrs={'class': 'webdirect-price d-series-ahb-switch'})
        tbody = tmp.find_parent('tbody')
        for row in tbody.find_all('tr'):
            data = row.find_all('td')
            if data[1].text.replace(' ', '').replace('\r\n', '').replace('\xa0', '') == 'D2v3':
                try:
                    price = float(json.loads(data[5].find('span')['data-amount'])['regional']['europe-west'])
                    self.service['2_8_windows']['price']['on_demand'] = price * 720 - self.service['ssd']['price']['on_demand'] * 50
                    price = float(json.loads(data[7].find('span')['data-amount'])['regional']['europe-west'])
                    self.service['2_8_windows']['price']['1_year'] = price * 720 - self.service['ssd']['price']['on_demand'] * 50
                    price = float(json.loads(data[9].find('span')['data-amount'])['regional']['europe-west'])
                    self.service['2_8_windows']['price']['3_year'] = price * 720 - self.service['ssd']['price']['on_demand'] * 50
                except Exception as err:
                    print('No data for Azure win D2 v3. ', err)

        # Standard Linux Configuration Prices
        req_st_lin_conf_prices = requests.get('https://azure.microsoft.com/ru-ru/pricing/details/virtual-machines/linux/',
                                              headers={'User-Agent': user_agent})
        soup_st_lin_conf_prices = BeautifulSoup(req_st_lin_conf_prices.text, "lxml")
        tmp = soup_st_lin_conf_prices.find('td', attrs={'class': 'webdirect-price av2-standard-series-ahb-switch'})
        tbody = tmp.find_parent('tbody')
        for row in tbody.find_all('tr'):
            data = row.find_all('td')
            if data[1].text.replace(' ', '').replace('\r\n', '').replace('\xa0', '') == 'A2v2':
                try:
                    price = json.loads(data[5].find('span')['data-amount'])['regional']['europe-west']
                    self.service['2_4_linux']['price']['on_demand'] = price * 720 - self.service['ssd']['price']['on_demand'] * 20
                except Exception as err:
                    print('No data for Azure lin A2 v2. ', err)

        tmp = soup_st_lin_conf_prices.find('td', attrs={'class': 'webdirect-price d-series-ahb-switch'})
        tbody = tmp.find_parent('tbody')
        for row in tbody.find_all('tr'):
            data = row.find_all('td')
            if data[1].text.replace(' ', '').replace('\r\n', '').replace('\xa0', '') == 'D2v3':
                try:
                    price = json.loads(data[5].find('span')['data-amount'])['regional']['europe-west']
                    self.service['2_8_linux']['price']['on_demand'] = price * 720 - self.service['ssd']['price']['on_demand'] * 50
                    price = json.loads(data[6].find('span')['data-amount'])['regional']['europe-west']
                    self.service['2_8_linux']['price']['1_year'] = price * 720 - self.service['ssd']['price']['on_demand'] * 50
                    price = json.loads(data[7].find('span')['data-amount'])['regional']['europe-west']
                    self.service['2_8_linux']['price']['3_year'] = price * 720 - self.service['ssd']['price']['on_demand'] * 50
                except Exception as err:
                    print('No data for Azure lin D2 v3. ', err)

        for commitment in self.service['2_4_80_linux']['price']:
            if self.service['2_4_linux']['price'][commitment] != None:
                self.service['2_4_80_linux']['price'][commitment] = self.service['2_4_linux']['price'][commitment] + self.service['ssd']['price']['on_demand'] * 80
            if self.service['2_4_windows']['price'][commitment] != None:
                self.service['2_4_80_windows']['price'][commitment] = self.service['2_4_windows']['price'][commitment] + self.service['ssd']['price']['on_demand'] * 80
            if self.service['2_8_linux']['price'][commitment] != None:
                self.service['2_8_160_linux']['price'][commitment] = self.service['2_8_linux']['price'][commitment] + self.service['ssd']['price']['on_demand'] * 160
            if self.service['2_8_windows']['price'][commitment] != None:
                self.service['2_8_160_windows']['price'][commitment] = self.service['2_8_windows']['price'][commitment] + self.service['ssd']['price']['on_demand'] * 160

    def get_GCP_prices(self):
        self.__init__()
        self.competitor_name = "GCP"
        self.region = 'Frankfurt'
        self.currency = 'USD'
        self.service['ssd']['comments'] = 'SSD provisioned space'
        self.service['ssd']['links'] = 'https://cloud.google.com/compute/disks-image-pricing'
        self.service['hdd']['comments'] = 'Standard provisioned space'
        self.service['hdd']['links'] = 'https://cloud.google.com/compute/disks-image-pricing'
        for service in ['cold_storage', 'cold_read', 'cold_write']:
            self.service[service]['comments'] = 'Nearline Storage'
            self.service[service]['links'] = 'https://cloud.google.com/storage/pricing'
        for service in ['standard_storage', 'standard_read', 'standard_write']:
            self.service[service]['comments'] = 'Standard Storage'
            self.service[service]['links'] = 'https://cloud.google.com/storage/pricing'
        for service in ['2_4_linux', '2_8_linux', '2_4_80_linux', '2_8_160_linux', '2_4_windows', '2_8_windows', '2_4_80_windows', '2_8_160_windows']:
            self.service[service]['comments'] = 'Custom machine type + Commited use discount'
            self.service[service]['links'] = 'https://cloud.google.com/compute/vm-instance-pricing'

        # SSD HDD Prices
        req_ssd_hdd = requests.get('https://cloud.google.com/compute/disks-image-pricing',
                                   headers={'User-Agent': user_agent})
        soup_ssd_hdd = BeautifulSoup(req_ssd_hdd.text, 'lxml')
        iframe_ssd_hdd = soup_ssd_hdd.find('iframe')
        req_ssd_hdd = requests.get("https://cloud-dot-devsite-v2-prod.appspot.com/" + iframe_ssd_hdd.attrs['src'],
                                   headers={'User-Agent': user_agent})
        soup_ssd_hdd = BeautifulSoup(req_ssd_hdd.text, 'lxml')
        tbody_ssd_hdd = soup_ssd_hdd.find_all('tbody')
        for tbody in tbody_ssd_hdd:
            for trow in tbody.find_all('tr'):
                    count = 0
                    flag = None
                    for tdata in trow.find_all('td'):
                        if count == 0:
                            if tdata.text == 'Standard provisioned space':
                                flag = tdata.text  # Standard provisioned space
                            if tdata.text == 'SSD provisioned space':
                                flag = tdata.text  # SSD provisioned space
                        if flag in {'Standard provisioned space', 'SSD provisioned space'} and count == 1:
                            price = float(tdata['ffurt-monthly'].replace('$', ''))
                            if flag == 'Standard provisioned space':
                                self.service['hdd']['price']['on_demand'] = price
                            if flag == 'SSD provisioned space':
                                self.service['ssd']['price']['on_demand'] = price
                        count += 1

        # Linux Configuration Prices
        Commitment = ['on_demand', '1_year', '3_year']
        Column = [1, 3, 4]
        vCPU_price = [0, 0, 0]
        RAM_price = [0, 0, 0]
        flag_cpu = 1
        flag_ram = 1
        req_st_lin_conf_prices = requests.get('https://cloud.google.com/compute/vm-instance-pricing',
                                              headers={'User-Agent': user_agent})
        soup_st_lin_conf_prices = BeautifulSoup(req_st_lin_conf_prices.text, "lxml")
        iframe_st_lin_conf_prices = soup_st_lin_conf_prices.find_all('iframe')
        for iframe in iframe_st_lin_conf_prices:
            req_st_lin_conf_prices = requests.get("https://cloud-dot-devsite-v2-prod.appspot.com/"
                                                  + iframe['src'],
                                                  headers={'User-Agent': user_agent})
            soup_st_lin_conf_prices = BeautifulSoup(req_st_lin_conf_prices.text, 'lxml')
            tbody_st_lin_conf_prices = soup_st_lin_conf_prices.find_all('tbody')
            for tbody in tbody_st_lin_conf_prices:
                for trow in tbody.find_all('tr'):
                    count = 0
                    flag = None
                    for tdata in trow.find_all('td'):
                        if count == 0:
                            if flag_cpu and tdata.text == 'Custom vCPUs':
                                flag = 'Custom vCPUs'
                                flag_cpu = 0
                        if flag in {'Custom vCPUs'} and count in Column:
                            vCPU_price[Column.index(count)] = \
                                float(tdata['ffurt-monthly'].split(' ')[0].replace('$', ''))
                        count += 1
            for tbody in tbody_st_lin_conf_prices:
                for trow in tbody.find_all('tr'):
                        count = 0
                        flag = None
                        for tdata in trow.find_all('td'):
                            if count == 0:
                                if flag_ram and tdata.text == 'Custom Memory':
                                    flag = 'Custom Memory'
                                    flag_ram = 0
                            if flag in {'Custom Memory'} and count in Column:
                                RAM_price[Column.index(count)] = \
                                    float(tdata['ffurt-monthly'].split(' ')[0].replace('$', ''))
                            count += 1
        for inc in range(3):
            if vCPU_price[inc] and RAM_price[inc]:
                self.service['cpu']['price'][Commitment[inc]] = vCPU_price[inc]
                self.service['ram']['price'][Commitment[inc]] = RAM_price[inc]
                self.service['2_4_linux']['price'][Commitment[inc]] = (2 * vCPU_price[inc] + 4 * RAM_price[inc])
                self.service['2_8_linux']['price'][Commitment[inc]] = (2 * vCPU_price[inc] + 8 * RAM_price[inc])
        # Windows Configuration Prices
        req_win_conf_prices = requests.get('https://cloud.google.com/compute/all-pricing',
                                              headers={'User-Agent': user_agent})
        soup_win_conf_prices = BeautifulSoup(req_win_conf_prices.text, "lxml")
        line_win_conf_prices = soup_win_conf_prices.find_all('li')
        for line in line_win_conf_prices:
            if 'for all other machine types' in ' '.join(line.text.split()):
                self.service['windows']['price']['on_demand'] = float(line.find('strong').text.split()[0].replace('$', '')) * 720
        for inc in range(3):
            if self.service['2_4_linux']['price'][Commitment[inc]] != None:
                self.service['2_4_windows']['price'][Commitment[inc]] = self.service['2_4_linux']['price'][Commitment[inc]] + \
                    + 2 * self.service['windows']['price']['on_demand']
            if self.service['2_8_linux']['price'][Commitment[inc]] != None:
                self.service['2_8_windows']['price'][Commitment[inc]] = self.service['2_8_linux']['price'][Commitment[inc]] + \
                    + 2 * self.service['windows']['price']['on_demand']

        # Cold Hot Storage Prices
        Type_Storage = ['Storage', 'Write', 'Read']
        GCP_name = {'Cold': 'Nearline Storage', 'Hot': 'Standard Storage'}
        req_cold_hot = requests.get('https://cloud.google.com/storage/pricing',
                                   headers={'User-Agent': user_agent})
        soup_cold_hot = BeautifulSoup(req_cold_hot.text, "lxml")
        # Read Write operations price extraction
        tbody_cold_hot = soup_cold_hot.find_all('html')
        for tbody in tbody_cold_hot:
            for trow in tbody.find_all('tr'):
                    count = 0
                    flag = None
                    for tdata in trow.find_all('td'):
                        if count == 0:
                            #print(tdata)
                            if tdata.find('strong'):
                                name = str(tdata.find('strong').text)
                                if name == GCP_name['Hot']:
                                    flag = name  # Standard Storage
                                if name == GCP_name['Cold']:
                                    flag = name  # Nearline Storage
                        if flag in {'Standard Storage', 'Nearline Storage'} and count in {1, 2}:
                            price = float(tdata.text.replace('$', ''))
                            if Type_Storage[count] == 'Write':
                                price = price/10
                            if flag == GCP_name['Hot']:
                                if Type_Storage[count] == 'Write':
                                    self.service['standard_write']['price']['on_demand'] = price
                                if Type_Storage[count] == 'Read':
                                    self.service['standard_read']['price']['on_demand'] = price
                            if flag == GCP_name['Cold']:
                                if Type_Storage[count] == 'Write':
                                    self.service['cold_write']['price']['on_demand'] = price
                                if Type_Storage[count] == 'Read':
                                    self.service['cold_read']['price']['on_demand'] = price
                        count += 1
        # Storage price extraction
        column_cold_hot = {'Hot': 0, 'Cold': 0}
        iframe_cold_hot = soup_cold_hot.find('iframe')
        req_cold_hot = requests.get("https://cloud-dot-devsite-v2-prod.appspot.com/"
                                    + iframe_cold_hot['src'],
                                    headers={'User-Agent': user_agent})
        soup_cold_hot = BeautifulSoup(req_cold_hot.text, "lxml")
        thead_cold_hot = soup_cold_hot.find('thead')
        count = 0
        for a in thead_cold_hot.find_all('a'):
            if a.text == GCP_name['Hot']:
                column_cold_hot['Hot'] = count
            if a.text == GCP_name['Cold']:
                column_cold_hot['Cold'] = count
            count += 1
        tbody_cold_hot = soup_cold_hot.find_all('tbody')
        for tbody in tbody_cold_hot:
            count = 0
            for tdata in tbody.find_all('td'):
                if count == column_cold_hot['Hot']:
                    self.service['standard_storage']['price']['on_demand'] = float(tdata['ffurt-monthly'].split(' ')[0].replace('$', ''))
                if count == column_cold_hot['Cold']:
                    self.service['cold_storage']['price']['on_demand'] = float(tdata['ffurt-monthly'].split(' ')[0].replace('$', ''))
                count += 1

        for commitment in self.service['2_4_80_linux']['price']:
            if self.service['2_4_linux']['price'][commitment] != None:
                self.service['2_4_80_linux']['price'][commitment] = self.service['2_4_linux']['price'][commitment] + self.service['ssd']['price']['on_demand'] * 80
            if self.service['2_4_windows']['price'][commitment] != None:
                self.service['2_4_80_windows']['price'][commitment] = self.service['2_4_windows']['price'][commitment] + self.service['ssd']['price']['on_demand'] * 80
            if self.service['2_8_linux']['price'][commitment] != None:
                self.service['2_8_160_linux']['price'][commitment] = self.service['2_8_linux']['price'][commitment] + self.service['ssd']['price']['on_demand'] * 160
            if self.service['2_8_windows']['price'][commitment] != None:
                self.service['2_8_160_windows']['price'][commitment] = self.service['2_8_windows']['price'][commitment] + self.service['ssd']['price']['on_demand'] * 160

    def get_AWS_prices(self):
        self.__init__()
        self.competitor_name = "AWS"
        self.region = 'Frankfurt'
        self.currency = 'USD'
        self.service['ssd']['comments'] = 'General Purpose'
        self.service['ssd']['links'] = 'https://aws.amazon.com/ebs/pricing/'
        self.service['hdd']['comments'] = 'Cold HDD'
        self.service['hdd']['links'] = 'https://aws.amazon.com/ebs/pricing/'
        for service in ['2_4_linux', '2_8_linux', '2_4_80_linux', '2_8_160_linux', '2_4_windows', '2_8_windows', '2_4_80_windows', '2_8_160_windows']:
            if '2_4' in service: self.service[service]['comments'] = 'c5.large'
            if '2_8' in service: self.service[service]['comments'] = 'c5.large'
            if 'windows' in service: self.service[service]['links'] = 'https://a0.p.awsstatic.com/pricing/1.0/ec2/region/eu-central-1/ondemand/windows/index.json'
            if 'windows' in service: self.service[service]['links'] = 'https://a0.p.awsstatic.com/pricing/1.0/ec2/region/eu-central-1/ondemand/linux/index.json'
        self.service['ssd']['price']['on_demand'] = 0.12
        self.service['hdd']['price']['on_demand'] = 0.03

        req_amazon = requests.get(
            'https://a0.p.awsstatic.com/pricing/1.0/ec2/region/eu-central-1/ondemand/linux/index.json',
            headers={'User-Agent': user_agent})
        soup_amazon = BeautifulSoup(req_amazon.text, "lxml")
        json_ec2_linux = json.loads(soup_amazon.find('p').text)['prices']
        for machine in json_ec2_linux:
            if machine['attributes']['aws:ec2:instanceType'] == 'c5.large':
                self.service['2_4_linux']['price']['on_demand'] = float(machine['price']['USD']) * 720 
            if machine['attributes']['aws:ec2:instanceType'] == 'm4.large':
                self.service['2_8_linux']['price']['on_demand'] = float(machine['price']['USD']) * 720 

        req_amazon = requests.get(
            'https://a0.p.awsstatic.com/pricing/1.0/ec2/region/eu-central-1/ondemand/windows/index.json',
            headers={'User-Agent': user_agent})
        soup_amazon = BeautifulSoup(req_amazon.text, "lxml")
        json_ec2_windows = json.loads(soup_amazon.find('p').text)['prices']
        for machine in json_ec2_windows:
            if machine['attributes']['aws:ec2:instanceType'] == 'c5.large':
                self.service['2_4_windows']['price']['on_demand'] = float(machine['price']['USD']) * 720 
            if machine['attributes']['aws:ec2:instanceType'] == 'm4.large':
                self.service['2_8_windows']['price']['on_demand'] = float(machine['price']['USD']) * 720 

        for commitment in self.service['2_4_80_linux']['price']:
            if self.service['2_4_linux']['price'][commitment] != None:
                self.service['2_4_80_linux']['price'][commitment] = self.service['2_4_linux']['price'][commitment] + self.service['ssd']['price']['on_demand'] * 80
            if self.service['2_4_windows']['price'][commitment] != None:
                self.service['2_4_80_windows']['price'][commitment] = self.service['2_4_windows']['price'][commitment] + self.service['ssd']['price']['on_demand'] * 80
            if self.service['2_8_linux']['price'][commitment] != None:
                self.service['2_8_160_linux']['price'][commitment] = self.service['2_8_linux']['price'][commitment] + self.service['ssd']['price']['on_demand'] * 160
            if self.service['2_8_windows']['price'][commitment] != None:
                self.service['2_8_160_windows']['price'][commitment] = self.service['2_8_windows']['price'][commitment] + self.service['ssd']['price']['on_demand'] * 160

    def get_DO_prices(self):
        self.competitor_name = 'DigitalOcean'
        self.currency = 'USD'
        self.region = ''
        for service in self.service:
            if '2_4' in service: self.service[service]['comments'] = 'CPU-Optimized Droplets 2 vCPU 4 RAM'
            if '2_8' in service: self.service[service]['comments'] = 'General Purpose Droplets 2 vCPU 8 RAM'
            self.service[service]['links'] = 'https://www.digitalocean.com/pricing/'
        self.service['ssd']['price']['on_demand'] = 0.1

        req_do = requests.get('https://www.digitalocean.com/pricing/', headers={'User-Agent': user_agent})
        soup_do = BeautifulSoup(req_do.text, "lxml")
        linux_8_2_25 = soup_do.find('a',
                                      attrs={'href': 'https://cloud.digitalocean.com/droplets/new?size=g-2vcpu-8gb'})
        linux_8_2_25_price = \
            float(linux_8_2_25.find_parent('tr').find_all('td')[4].text.replace('$', '')) * 720
        self.service['2_8_linux']['price']['on_demand'] = linux_8_2_25_price - 25 * self.service['ssd']['price']['on_demand']

        linux_4_2_25 = soup_do.find('a', attrs={'href': 'https://cloud.digitalocean.com/droplets/new?size=c-2-4gib'})
        linux_4_2_25_price = \
            float(linux_4_2_25.find_parent('tr').find_all('td')[4].text.replace('$', '')) * 720
        self.service['2_4_linux']['price']['on_demand'] = linux_4_2_25_price - 25 * self.service['ssd']['price']['on_demand']

        object_storage = soup_do.find('a', attrs={'href': 'https://cloud.digitalocean.com/registrations/'
                                                          'new?onboarding_origin=spaces'})
        object_storage_price = \
            float(object_storage.find_parent('tr').find_all('td')[2].text.replace('$', '').replace('/GB', ''))
        self.service['cold_storage']['price']['on_demand'] = object_storage_price
        self.service['standard_storage']['price']['on_demand'] = object_storage_price
        self.service['cold_read']['price']['on_demand'] = 0
        self.service['standard_read']['price']['on_demand'] = 0
        self.service['cold_write']['price']['on_demand'] = 0
        self.service['standard_write']['price']['on_demand'] = 0

        for commitment in self.service['2_4_80_linux']['price']:
            if self.service['2_4_linux']['price'][commitment] != None:
                self.service['2_4_80_linux']['price'][commitment] = self.service['2_4_linux']['price'][commitment] + self.service['ssd']['price']['on_demand'] * 80
            if self.service['2_4_windows']['price'][commitment] != None:
                self.service['2_4_80_windows']['price'][commitment] = self.service['2_4_windows']['price'][commitment] + self.service['ssd']['price']['on_demand'] * 80
            if self.service['2_8_linux']['price'][commitment] != None:
                self.service['2_8_160_linux']['price'][commitment] = self.service['2_8_linux']['price'][commitment] + self.service['ssd']['price']['on_demand'] * 160
            if self.service['2_8_windows']['price'][commitment] != None:
                self.service['2_8_160_windows']['price'][commitment] = self.service['2_8_windows']['price'][commitment] + self.service['ssd']['price']['on_demand'] * 160

    def get_Selectel_prices(self):
        self.competitor_name = 'Selectel'
        self.currency = 'RUB'
        self.region = 'ru-7'
        self.service['ssd']['comments'] = 'Универсальный диск'
        self.service['hdd']['comments'] = 'Базовый диск'
        for service in self.service:
            if '2_' in service: self.service[service]['comments'] = 'Custom machine type'
            if ('cold' in service) or ('standart' in service):
                self.service[service]['links'] = 'https://selectel.ru/services/cloud/storage/#prices'
            else: self.service[service]['links'] = 'https://selectel.ru/services/cloud/servers/prices/'

        cpu_price = 0
        ram_price = 0
        ssd_price = 0
        hdd_price = 0
        win_price = 0
        req = requests.get("https://selectel.ru/api/prices/vpc?currency=rub",
                           headers={'User-Agent': user_agent})
        req.encoding = req.apparent_encoding
        soup = BeautifulSoup(req.text, "lxml")
        table = soup.find('p')
        for price in json.loads(table.text):
            if re.match("ru-7.*", price['group']):
                if price['resource'] == 'compute_cores':
                    cpu_price = price['value'] * 720 / 100 / (1 + VAT)
                if price['resource'] == 'compute_ram':
                    ram_price = price['value'] * 1024 * 720 / 100 / (1 + VAT)
                if price['resource'] == 'volume_gigabytes_universal':
                    ssd_price = price['value'] * 720 / 100 / (1 + VAT)
                if price['resource'] == 'volume_gigabytes_basic':
                    hdd_price = price['value'] * 720 / 100 / (1 + VAT)
                if price['resource'] == 'license_windows_2019_standard':
                    win_price = price['value'] * 720 / 100 / (1 + VAT)
        self.service['cpu']['price']['on_demand'] = cpu_price
        self.service['ram']['price']['on_demand'] = ram_price
        self.service['windows']['price']['on_demand'] = win_price
        self.service['2_4_windows']['price']['on_demand'] =  2 * cpu_price + 4 * ram_price + win_price
        self.service['2_8_windows']['price']['on_demand'] =  2 * cpu_price + 8 * ram_price + win_price
        self.service['2_4_linux']['price']['on_demand'] = 2 * cpu_price + 4 * ram_price
        self.service['ssd']['price']['on_demand'] = ssd_price
        self.service['hdd']['price']['on_demand'] = hdd_price

        req = requests.get("https://selectel.ru/api/prices/storage?currency=rub",
                           headers={'User-Agent': user_agent})
        req.encoding = req.apparent_encoding
        soup = BeautifulSoup(req.text, "lxml")
        table = soup.find('p')
        for price in json.loads(table.text):
            if price['resource'] == 'space-mb' and price['threshold'] == 1048576:                    
                self.service['cold_storage']['price']['on_demand'] = price['value'] * 1024 * 720 / 100 / (1 + VAT)
                self.service['standard_storage']['price']['on_demand'] = self.service['cold_storage']['price']['on_demand']
            if price['resource'] == 'storage-traffic-req-put':
                self.service['cold_write']['price']['on_demand'] = price['value'] * 1000 / 100 / (1 + VAT)
                self.service['standard_write']['price']['on_demand'] = self.service['cold_write']['price']['on_demand']
            if price['resource'] == 'storage-traffic-req-get':
                self.service['cold_read']['price']['on_demand'] = price['value'] * 10000 / 100 / (1 + VAT)
                self.service['standard_read']['price']['on_demand'] = self.service['cold_read']['price']['on_demand']

        for commitment in self.service['2_4_80_linux']['price']:
            if self.service['2_4_linux']['price'][commitment] != None:
                self.service['2_4_80_linux']['price'][commitment] = self.service['2_4_linux']['price'][commitment] + self.service['ssd']['price']['on_demand'] * 80
            if self.service['2_4_windows']['price'][commitment] != None:
                self.service['2_4_80_windows']['price'][commitment] = self.service['2_4_windows']['price'][commitment] + self.service['ssd']['price']['on_demand'] * 80
            if self.service['2_8_linux']['price'][commitment] != None:
                self.service['2_8_160_linux']['price'][commitment] = self.service['2_8_linux']['price'][commitment] + self.service['ssd']['price']['on_demand'] * 160
            if self.service['2_8_windows']['price'][commitment] != None:
                self.service['2_8_160_windows']['price'][commitment] = self.service['2_8_windows']['price'][commitment] + self.service['ssd']['price']['on_demand'] * 160

    def get_MCS_prices(self):
        self.competitor_name = 'MCS'
        self.currency = 'RUB'
        self.region = ''
        for service in self.service:
            if '2_' in service: self.service[service]['comments'] = 'Custom machine type'
            if ('cold' in service) or ('standart' in service):
                self.service[service]['links'] = 'https://mcs.mail.ru/help/storage-pay/storage-tariffs'
            else:
                self.service[service]['links'] = 'https://mcs.mail.ru/help/iaas-pay/tariffs'
        
        cpu_price = 0
        ram_price = 0
        ssd_price = 0
        hdd_price = 0
        win_price = 315
        req = requests.get("https://mcs.mail.ru/help/iaas-pay/tariffs",
                           headers={'User-Agent': user_agent})
        req.encoding = req.apparent_encoding
        soup = BeautifulSoup(req.text, "lxml")
        for text in soup.find_all('span', attrs={'style': "color: #888888;"}):
            for part_text in text.find('strong').text.replace(' ', '').replace('\xa0', '').replace('₽/мес.', '').split(','):
                nameprice = part_text.split('—')
                if nameprice[0] == '1ГБпамяти':
                    ram_price = float(nameprice[1]) / (1 + VAT)
                if nameprice[0] == '1CPU':
                    cpu_price = float(nameprice[1]) / (1 + VAT)
                if nameprice[0] == '1ГБHDD-диска':
                    hdd_price = float(nameprice[1]) / (1 + VAT)
                if nameprice[0] == '1ГБSSDдиска':
                    ssd_price = float(nameprice[1]) / (1 + VAT)
        self.service['cpu']['price']['on_demand'] = cpu_price
        self.service['ram']['price']['on_demand'] = ram_price
        self.service['windows']['price']['on_demand'] = win_price
        self.service['2_4_windows']['price']['on_demand'] =  2 * cpu_price + 4 * ram_price + win_price
        self.service['2_8_windows']['price']['on_demand'] =  2 * cpu_price + 8 * ram_price + win_price
        self.service['2_4_linux']['price']['on_demand'] = 2 * cpu_price + 4 * ram_price
        self.service['2_8_linux']['price']['on_demand'] = 2 * cpu_price + 8 * ram_price
        self.service['ssd']['price']['on_demand'] = ssd_price
        self.service['hdd']['price']['on_demand'] = hdd_price

        req = requests.get("https://mcs.mail.ru/help/storage-pay/storage-tariffs",
                           headers={'User-Agent': user_agent})
        req.encoding = req.apparent_encoding
        soup = BeautifulSoup(req.text, "lxml")
        inc = 0
        for text in soup.find_all('td', attrs={'style': "width: 23.5417%;"}):
            inc += 1
            price = 0
            try:
                price = float(re.match('.*руб', text.text.replace(',', '.')).group().replace(' руб', ''))
            except:
                a = 0
            if inc == 2:
                self.service['standard_storage']['price']['on_demand'] = price / (1 + VAT)
            if inc == 5:
                self.service['standard_write']['price']['on_demand'] = price / (1 + VAT)
            if inc == 6:
                self.service['standard_read']['price']['on_demand'] = price / (1 + VAT)
        inc = 0
        for text in soup.find_all('td', attrs={'style': "width: 22.7083%;"}):
            inc += 1
            price = 0
            try:
                price = float(re.match('.*руб', text.text.replace(',', '.')).group().replace(' руб', ''))
            except:
                a = 0
            if inc == 2:
                self.service['cold_storage']['price']['on_demand'] = price / (1 + VAT)
            if inc == 5:
                self.service['cold_write']['price']['on_demand'] = price / (1 + VAT)
            if inc == 6:
                self.service['cold_read']['price']['on_demand'] = price / (1 + VAT)

        for commitment in self.service['2_4_80_linux']['price']:
            if self.service['2_4_linux']['price'][commitment] != None:
                self.service['2_4_80_linux']['price'][commitment] = self.service['2_4_linux']['price'][commitment] + self.service['ssd']['price']['on_demand'] * 80
            if self.service['2_4_windows']['price'][commitment] != None:
                self.service['2_4_80_windows']['price'][commitment] = self.service['2_4_windows']['price'][commitment] + self.service['ssd']['price']['on_demand'] * 80
            if self.service['2_8_linux']['price'][commitment] != None:
                self.service['2_8_160_linux']['price'][commitment] = self.service['2_8_linux']['price'][commitment] + self.service['ssd']['price']['on_demand'] * 160
            if self.service['2_8_windows']['price'][commitment] != None:
                self.service['2_8_160_windows']['price'][commitment] = self.service['2_8_windows']['price'][commitment] + self.service['ssd']['price']['on_demand'] * 160

    def get_usd(self, type, azure_e='default'):
        exch_rate = 0
        if type == 'Azure':
            req = requests.get("https://functionaed97c258575.blob.core.windows.net/azureexchrates/lastexchrates.json",
                            headers={'User-Agent': user_agent})
            soup = BeautifulSoup(req.text, "lxml")
            for tmp in json.loads(soup.find('p').text)['ExchRates']:
                if tmp['Currency'] == 'RUB':
                    exch_rate = tmp['AzureEchangeRateToUSD']
                    print(azure_e, exch_rate)          
            if azure_e == 'default':
                exch_rate = 62.5 #Solid as for 20200603 there was no information on site
        else:
            req = requests.get("https://www.google.com/search?source="
                            "hp&ei=YmqbXZClBu-rrgSv7KDoDQ&q=USD+RUB&oq=USD+RUB"
                            "&gs_l=psy-ab.3...0.0..2389...0.0..0.0.0.......0......gws-wiz."
                            "&ved=0ahUKEwiQ3p2gy4rlAhXvlYsKHS82CN0Q4dUDCAY&uact=5",
                            headers={'User-Agent': 'Mozilla/5.0'})
            soup = BeautifulSoup(req.text, "lxml")
            exch_rate = soup.find_all('div', attrs={'class':'BNeawe iBp4i AP7Wnd'})[0].text.rsplit(' ', 2)[0].replace(',', '.')
        return float(exch_rate)

    def print(self):
        for service in self.service:
            for commitment in self.service[service]['price']:
                print(service, commitment, self.service[service]['price'][commitment])

    def append_df(self, df_prices):
        cur_day = pd.Timestamp(time.strftime('%Y%m%d', time.localtime()))
        df_tmp = pd.DataFrame(columns=df_prices.columns)
        try:
            df_tmp = pd.DataFrame(columns=df_prices.columns)
            inc = 0
            for service in self.service:
                for commitment in self.service[service]['price']:
                    df_tmp.loc[inc] = [cur_day, self.competitor_name, service, 
                                    commitment, self.service[service]['price'][commitment], self.service[service]['pricing_unit'], 
                                    self.currency, self.service[service]['links'], self.service[service]['comments'], self.region]
                    inc += 1

            for index, row in df_tmp.iterrows():
                if row['price'] == None:
                    df_tmp.drop(index, inplace=True)
            df_tmp = df_tmp.dropna()

            usd_rub = self.get_usd(type=self.competitor_name)
            if self.currency == 'USD':
                df_tmp_conv = df_tmp.copy()
                df_tmp_conv['price'] = df_tmp_conv['price'] * usd_rub
                df_tmp_conv['currency'] = ['RUB'] * len(df_tmp_conv['currency'])
                df_tmp = df_tmp.append(df_tmp_conv, ignore_index=True)
            elif self.currency == 'RUB':
                df_tmp_conv = df_tmp.copy()
                df_tmp_conv['price'] = df_tmp_conv['price'] / usd_rub
                df_tmp_conv['currency'] = ['USD'] * len(df_tmp_conv['currency'])
                df_tmp = df_tmp.append(df_tmp_conv, ignore_index=True)

        except Exception as err:
            print(self.competitor_name, 'problems with append', err)
    
        df_prices = df_prices.append(df_tmp, ignore_index=True)
        return df_prices
    
    def export(self, path, df):
        
        cur_day = time.strftime('%Y%m%d', time.localtime())
        #To Excel
        try:
            data = pd.read_excel('{}.xlsx'.format(path), sheet_name="Prices")

                # data.to_excel('{}_{}.xlsx'.format(path, cur_day), sheet_name='Prices')
                # writer = pd.ExcelWriter('{}_{}.xlsx'.format(path, cur_day), engine='openpyxl')
                # book = load_workbook('{}.xlsx'.format(path))
                # writer.book = book
                # writer.sheets = {ws.title: ws for ws in book.worksheets}
                # df.to_excel(writer, sheet_name='Prices', startrow=writer.sheets['Prices'].max_row, index=False, header=False)
                # writer.save()
            
            usd_rub = self.get_usd(type='AWS')
            data_usd = data.loc[(data['currency'] == 'USD') & (data['company'] != 'YandexCloud')].copy()
            data_usd['price'] = data_usd['price'] * usd_rub
            data_usd['currency'] = ['RUB'] * len(data_usd['currency'])

            data_rub = data.loc[(data['currency'] == 'RUB') & (data['company'] != 'YandexCloud')].copy()
            data_rub['price'] = data_rub['price'] / usd_rub
            data_rub['currency'] = ['USD'] * len(data_rub['currency'])
            
            data = data.append(data_usd, ignore_index=True)
            data = data.append(data_rub, ignore_index=True)

            df = df.append(data, ignore_index=True)
            df.to_excel('{}_{}.xlsx'.format(path, cur_day), sheet_name='Prices', index=False, header=False)
            print('Data was extracted to Excel')
            df.to_csv(path, index=False, header=False)
            print('Data was extracted to CSV')
        except Exception as err:
            print('Problems with data extraction to Excel, it can affect to Hahan extraction: ', err)

        #To YT
        try:            
            df['update'] = df['update'].apply(lambda x: (x - pd.Timestamp('1970-01-01')) // pd.Timedelta('1s'))
            df['comments'] = df['comments'].apply(lambda x: str(x))
            df['region'] = df['region'].apply(lambda x: str(x))

            table = "//home/cloud_analytics/dkharitonov/competitors_prices/competitors_prices"            
            cur_table = table + "_" + cur_day
            yt.wrapper.config.set_proxy("hahn")
            schema = [{"name" : "update", "type" : "datetime"}, 
                        {"name" : "company", "type" : "string"}, 
                        {"name" : "service", "type" : "string"}, 
                        {"name" : "commitment", "type" : "string"}, 
                        {"name" : "price", "type" : "double"}, 
                        {"name" : "pricing_unit", "type" : "string"}, 
                        {"name" : "currency", "type" : "string"}, 
                        {"name" : "link", "type" : "string"},
                        {"name" : "comments", "type" : "string"},
                        {"name" : "region", "type" : "string"}
                     ]
            yt.wrapper.create('table', path=cur_table, attributes={"schema" : schema})
            yt.wrapper.write_table(cur_table, df.to_dict(orient='records'))
            yt.wrapper.remove(table)
            yt.wrapper.copy(cur_table, table)
            print('Data was extracted to Hahn: ', table)
        except Exception as err:
            print('Problems with data extraction to Hahn: ', err)


def main():
    col = ['update', 'company', 'service', 'commitment', 'price', 'pricing_unit', 'currency', 'link', 'comments', 'region']
    df_prices = pd.DataFrame(columns=col)

    prices = GetPrices()

    try:
        prices = GetPrices()
        prices.get_Azure_prices()
        df_prices = prices.append_df(df_prices)
    except Exception as err:
        print('Problems with Azure extraction\n', err)

    try:
        prices = GetPrices()
        prices.get_GCP_prices()
        df_prices = prices.append_df(df_prices)
    except Exception as err:
        print('Problems with GCP extraction\n', err)

    try:
        prices = GetPrices()
        prices.get_AWS_prices()
        df_prices = prices.append_df(df_prices)
    except Exception as err:
        print('Problems with AWS extraction\n', err)

    try:
        prices = GetPrices()
        prices.get_DO_prices()
        df_prices = prices.append_df(df_prices)
    except Exception as err:
        print('Problems with DO extraction\n', err)

    try:
        prices = GetPrices()
        prices.get_Selectel_prices()
        df_prices = prices.append_df(df_prices)
    except Exception as err:
        print('Problems with Selectel extraction\n', err)

    try:
        prices = GetPrices()
        prices.get_MCS_prices()
        df_prices = prices.append_df(df_prices)
    except Exception as err:
        print('Problems with MCS extraction\n', err)

    prices.export(path='/home/daniilkhar/competitors_prices/competitors_prices', df=df_prices)

    print('USD exchange rate for Azure from site: ', prices.get_usd(type='Azure', azure_e='real'),
          '\nUSD exchange rate for Azure by default: ', prices.get_usd(type='Azure'), 
          '\nUSD exchange rate : ', prices.get_usd(type=''))


if __name__=='__main__':
    main()
